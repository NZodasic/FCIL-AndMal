"""Tests for the complete classification metric and result-reporting schema."""

import importlib.util
import csv
import tempfile
import unittest
from pathlib import Path

import numpy as np

from utils.logging import ExperimentLogger
from utils.metrics import (
    CLASSIFICATION_METRIC_KEYS,
    compute_classification_metrics,
    format_classification_metrics,
    format_confusion_matrix,
)
from utils.results import (
    RESULT_COLUMNS,
    build_result_rows,
    final_round_for_task,
    save_task_confusion_matrix,
    write_experiment_sheet,
)


OPENPYXL_AVAILABLE = importlib.util.find_spec("openpyxl") is not None


class TestMetricsReporting(unittest.TestCase):
    def test_all_ten_classification_metrics_are_reported(self):
        metrics = compute_classification_metrics(
            np.array([0, 0, 1, 1, 2, 2]),
            np.array([0, 1, 1, 1, 2, 0]),
            seen_classes=[0, 1, 2],
        )

        expected = {
            "accuracy",
            "precision_macro",
            "precision_micro",
            "precision_weighted",
            "recall_macro",
            "recall_micro",
            "recall_weighted",
            "f1_macro",
            "f1_micro",
            "f1_weighted",
        }
        self.assertTrue(expected.issubset(metrics))
        self.assertEqual(len(metrics["confusion_matrix"]), 3)
        self.assertEqual(metrics["confusion_matrix_labels"], [0, 1, 2])

        formatted = format_classification_metrics(metrics)
        for label in (
            "Accuracy",
            "Macro Precision",
            "Macro Recall",
            "Macro F1",
            "Micro Precision",
            "Micro Recall",
            "Micro F1",
            "Weighted Precision",
            "Weighted Recall",
            "Weighted F1",
        ):
            self.assertIn(label, formatted)

    def test_evaluation_logger_records_all_metrics_and_final_matrix(self):
        metrics = compute_classification_metrics(
            np.array([0, 0, 1, 1]),
            np.array([0, 1, 1, 1]),
            seen_classes=[0, 1],
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            logger = ExperimentLogger(temp_dir, "reporting")
            logger.log_evaluation(
                metrics,
                context="Task 1 | Epoch 5/10 | Test",
                task_id=0,
                step=5,
            )
            intermediate_log = Path(temp_dir, "reporting.log").read_text()
            self.assertNotIn("Confusion Matrix", intermediate_log)

            logger.log_evaluation(
                metrics,
                context="Task 1 | Epoch 10/10 | Final Test",
                task_id=0,
                step=10,
                include_confusion_matrix=True,
                label_names={0: "Benign", 1: "Malware"},
            )
            logger.close()

            log_text = Path(temp_dir, "reporting.log").read_text()
            self.assertIn("Confusion Matrix", log_text)
            self.assertIn("Benign", format_confusion_matrix(
                metrics, label_names={0: "Benign", 1: "Malware"}
            ))
            with Path(temp_dir, "reporting_metrics.csv").open(newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(len(rows), 2)
            self.assertTrue(set(CLASSIFICATION_METRIC_KEYS).issubset(rows[0]))

    def test_federated_rows_use_cumulative_task_final_rounds(self):
        task_results = []
        for task_id in range(5):
            result = {column: 0.5 for column in RESULT_COLUMNS[8:18]}
            result["task_id"] = task_id
            task_results.append(result)

        rows = build_result_rows(
            task_results,
            experiment_name="case",
            method="malfsil",
            setting="federated",
            rounds_per_task=50,
            client_num=20,
            patch_size=256,
            test_location="test.parquet",
            confusion_matrix_paths=[f"cm_{i}.png" for i in range(5)],
            checkpoint_paths=[f"checkpoint_{i}.pt" for i in range(5)],
        )

        self.assertEqual([row["round"] for row in rows], [50, 100, 150, 200, 250])
        self.assertTrue(all(list(row) == RESULT_COLUMNS for row in rows))

    def test_centralized_rows_use_final_epoch(self):
        metrics = {column: 1.0 for column in RESULT_COLUMNS[8:18]}
        metrics["task_id"] = 0
        row = build_result_rows(
            [metrics],
            experiment_name="central",
            method="joint",
            setting="centralized",
            rounds_per_task=25,
            client_num="N/A",
            patch_size=128,
            test_location="test.csv",
            confusion_matrix_paths=["cm.png"],
            checkpoint_paths=["checkpoint.pt"],
        )[0]
        self.assertEqual(row["round"], 25)
        self.assertEqual(row["task_id"], 1)

    def test_final_round_validation(self):
        self.assertEqual(final_round_for_task(4, 50), 250)
        with self.assertRaises(ValueError):
            final_round_for_task(-1, 50)

    def test_confusion_matrix_artifact_is_created(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = save_task_confusion_matrix(
                {
                    "confusion_matrix": [[2, 1], [0, 3]],
                    "confusion_matrix_labels": [0, 1],
                },
                task_id=0,
                output_dir=temp_dir,
            )
            self.assertTrue(Path(path).is_file())

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
    def test_workbook_preserves_one_sheet_per_experiment(self):
        from openpyxl import load_workbook

        row = {column: "value" for column in RESULT_COLUMNS}
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = str(Path(temp_dir) / "results.xlsx")
            write_experiment_sheet(workbook_path, "case_a", [row])
            write_experiment_sheet(workbook_path, "case_b", [row])
            write_experiment_sheet(workbook_path, "case_a", [row, row])

            workbook = load_workbook(workbook_path, read_only=True)
            self.assertEqual(set(workbook.sheetnames), {"case_a", "case_b"})
            self.assertEqual(
                [cell.value for cell in workbook["case_a"][1]], RESULT_COLUMNS
            )
            self.assertEqual(workbook["case_a"].max_row, 3)


if __name__ == "__main__":
    unittest.main()
