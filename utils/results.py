"""Excel reporting for centralized and federated experiment results."""

from hashlib import sha1
from pathlib import Path
from typing import Any, Dict, List, Sequence, Union

from config import ID2LABEL
from utils.plotting import plot_confusion_matrix


RESULT_COLUMNS = [
    "text",
    "method",
    "setting",
    "task_id",
    "round",
    "client_num",
    "patch_size",
    "test_location",
    "accuracy",
    "precision_macro",
    "precision_micro",
    "precision_weighted",
    "recall_macro",
    "recall_micro",
    "recall_weighted",
    "f1_macro",
    "f1_micro",
    "f1_weighted",
    "confusion_matrix_path",
    "checkpoint_path",
]

METRIC_COLUMNS = RESULT_COLUMNS[8:18]


def resolve_test_location(prepared_data_dir: str, feature_type: str) -> str:
    """Return the held-out test file actually used by the evaluator."""
    test_dir = Path(prepared_data_dir) / feature_type
    for filename in ("test.parquet", "test.csv"):
        candidate = test_dir / filename
        if candidate.is_file():
            return str(candidate.resolve())
    return str((test_dir / "test.parquet").resolve())


def final_round_for_task(task_id: int, rounds_per_task: int) -> int:
    """Convert a zero-based task ID to its cumulative final global round."""
    if task_id < 0:
        raise ValueError("task_id must be non-negative")
    if rounds_per_task <= 0:
        raise ValueError("rounds_per_task must be positive")
    return (task_id + 1) * rounds_per_task


def build_result_rows(
    task_results: Sequence[Dict[str, Any]],
    *,
    experiment_name: str,
    method: str,
    setting: str,
    rounds_per_task: int,
    client_num: Union[int, str],
    patch_size: int,
    test_location: str,
    confusion_matrix_paths: Sequence[str],
    checkpoint_paths: Sequence[str],
) -> List[Dict[str, Any]]:
    """Build workbook rows from task-final evaluation dictionaries."""
    if not (
        len(task_results) == len(confusion_matrix_paths) == len(checkpoint_paths)
    ):
        raise ValueError("Every task result needs a confusion matrix and checkpoint path")

    normalized_setting = setting.lower()
    if normalized_setting not in {"centralized", "federated"}:
        raise ValueError("setting must be 'centralized' or 'federated'")

    rows = []
    for index, metrics in enumerate(task_results):
        task_id = int(metrics.get("task_id", index))
        final_round = (
            final_round_for_task(task_id, rounds_per_task)
            if normalized_setting == "federated"
            else rounds_per_task
        )
        row = {
            "text": experiment_name,
            "method": method,
            "setting": normalized_setting,
            "task_id": task_id + 1,
            "round": final_round,
            "client_num": client_num,
            "patch_size": patch_size,
            "test_location": str(Path(test_location).resolve()),
            "confusion_matrix_path": str(Path(confusion_matrix_paths[index]).resolve()),
            "checkpoint_path": str(Path(checkpoint_paths[index]).resolve()),
        }
        for column in METRIC_COLUMNS:
            row[column] = float(metrics.get(column, 0.0))
        rows.append({column: row[column] for column in RESULT_COLUMNS})
    return rows


def save_task_confusion_matrix(
    metrics: Dict[str, Any],
    task_id: int,
    output_dir: str,
) -> str:
    """Render the confusion matrix embedded in a task evaluation result."""
    matrix = metrics.get("confusion_matrix", [])
    labels = metrics.get("confusion_matrix_labels", [])
    if not matrix or not labels:
        raise ValueError(f"Task {task_id + 1} has no confusion matrix data")

    class_names = [ID2LABEL.get(int(label), str(label)) for label in labels]
    output_path = Path(output_dir) / f"task_{task_id + 1}_confusion_matrix.png"
    plot_confusion_matrix(
        matrix,
        class_names,
        str(output_path),
        title=f"Task {task_id + 1} Confusion Matrix",
    )
    return str(output_path.resolve())


def _sheet_title(experiment_name: str) -> str:
    invalid = set("[]:*?/\\")
    cleaned = "".join("_" if char in invalid else char for char in experiment_name).strip()
    cleaned = cleaned or "experiment"
    if len(cleaned) <= 31:
        return cleaned
    digest = sha1(experiment_name.encode("utf-8")).hexdigest()[:7]
    return f"{cleaned[:23]}_{digest}"


def _excel_value(value: Any) -> Any:
    # Prevent user-controlled experiment names from becoming spreadsheet formulas.
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def write_experiment_sheet(
    workbook_path: str,
    experiment_name: str,
    rows: Sequence[Dict[str, Any]],
) -> str:
    """Create or replace one experiment sheet without changing other sheets."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError(
            "Excel export requires openpyxl; install the project requirements"
        ) from exc

    output_path = Path(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(output_path) if output_path.exists() else Workbook()
    title = _sheet_title(experiment_name)

    if title in workbook.sheetnames:
        workbook.remove(workbook[title])
    sheet = workbook.create_sheet(title)
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        workbook.remove(workbook["Sheet"])

    sheet.append(RESULT_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:T{max(1, len(rows) + 1)}"

    for row in rows:
        sheet.append([_excel_value(row.get(column, "")) for column in RESULT_COLUMNS])

    for column_cells in sheet.columns:
        width = min(60, max(len(str(cell.value or "")) for cell in column_cells) + 2)
        sheet.column_dimensions[column_cells[0].column_letter].width = width

    workbook.save(output_path)
    return str(output_path.resolve())


def export_experiment_results(
    *,
    workbook_path: str,
    experiment_name: str,
    method: str,
    setting: str,
    rounds_per_task: int,
    client_num: Union[int, str],
    patch_size: int,
    test_location: str,
    task_results: Sequence[Dict[str, Any]],
    checkpoint_paths: Sequence[str],
    artifact_dir: str,
) -> str:
    """Save task confusion matrices and the experiment's Excel worksheet."""
    confusion_paths = [
        save_task_confusion_matrix(metrics, index, artifact_dir)
        for index, metrics in enumerate(task_results)
    ]
    rows = build_result_rows(
        task_results,
        experiment_name=experiment_name,
        method=method,
        setting=setting,
        rounds_per_task=rounds_per_task,
        client_num=client_num,
        patch_size=patch_size,
        test_location=test_location,
        confusion_matrix_paths=confusion_paths,
        checkpoint_paths=checkpoint_paths,
    )
    return write_experiment_sheet(workbook_path, experiment_name, rows)
